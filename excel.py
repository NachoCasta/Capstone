from openpyxl import Workbook, load_workbook

def excel_to_csv(file):
    print("Comienza creación de csv's")
    wb = load_workbook(file)
    print("Excel abierto")
    hojas = [
        ("Ciudad", 2, 6),
        ("Supermercado", 2, 2),
        ("Locales", 1, 3),
        ("Histórico de Compras", 2, 12),
        ("Sensibilidad", 1, 1)
        ]
    for hoja, header, lf in hojas:
        print("Leyendo {}".format(hoja))
        sheet = wb[hoja]
        if hoja == "Ciudad":
            with open("bdd/Largo calles.csv", "w") as f:
                f.write("Largo de las calles [metros]\n")
                f.write(str(sheet["I3"].value))
        tabla = [[c.value for c in fila[1:lf+1]]
                 for fila in sheet.iter_rows(row_offset=header)]
        print("Creando {}.csv".format(hoja))
        with open("bdd/{}.csv".format(hoja), "w") as f:
            for fila in tabla:
                if hoja != "Sensibilidad":
                    if str(fila[1]) == "None":
                        break
                f.write(";".join(map(str, fila))+"\n")
        print("Creado {}.csv".format(hoja))

               
if __name__ == "__main__":
    excel_to_csv("bdd/Datos.xlsx")
